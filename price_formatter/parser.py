"""
Модуль загрузки и парсинга Excel-файлов.
Обрабатывает объединенные ячейки, пустые строки, шапки таблиц и разрывы страниц.
"""
import pandas as pd
from openpyxl import load_workbook
from typing import Tuple, List, Optional
import re


def load_excel_file(file_path: str) -> pd.DataFrame:
    """
    Загружает Excel файл с обработкой объединенных ячеек.
    """
    # Сначала читаем через openpyxl для обработки объединенных ячеек
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Распаковываем объединенные ячейки
    for merged in list(ws.merged_cells.ranges):
        top_left = merged.start_cell.value
        for row in range(merged.min_row, merged.max_row + 1):
            for col in range(merged.min_col, merged.max_col + 1):
                ws.cell(row=row, column=col).value = top_left
    
    # Сохраняем во временный файл для чтения pandas
    temp_path = file_path.replace('.xlsx', '_temp.xlsx')
    wb.save(temp_path)
    
    # Читаем через pandas
    df = pd.read_excel(temp_path, engine='openpyxl', header=None)
    
    import os
    os.remove(temp_path)
    
    return df


def find_table_start(df: pd.DataFrame, expected_headers: List[str]) -> int:
    """
    Находит строку начала таблицы данных по заголовкам.
    Возвращает индекс строки с заголовками.
    """
    for idx, row in df.iterrows():
        row_values = [str(v).strip() if pd.notna(v) else '' for v in row.values]
        matches = sum(1 for header in expected_headers 
                     if any(header.lower() in str(v).lower() for v in row_values))
        if matches >= 2:  # Найдено хотя бы 2 совпадения с ожидаемыми заголовками
            return idx
    
    return 0  # Если не найдено, начинаем с первой строки


def detect_columns(df: pd.DataFrame, header_row: int) -> dict:
    """
    Автоматически определяет индексы колонок по заголовкам.
    """
    headers = df.iloc[header_row].tolist()
    columns_map = {}
    
    header_keywords = {
        'code': ['код', 'код товара', 'артикул', 'id'],
        'status': ['статус', 'акция', 'распродажа'],
        'name': ['номенклатура', 'название', 'товар', 'наименование'],
        'special_price': ['спец', 'акция', 'специальная', 'новая цена'],
        'retail_price': ['розничная', 'базовая', 'цена', 'старая цена'],
    }
    
    for col_idx, header in enumerate(headers):
        if pd.isna(header):
            continue
        header_str = str(header).lower().strip()
        
        for col_type, keywords in header_keywords.items():
            for keyword in keywords:
                if keyword in header_str:
                    columns_map[col_type] = col_idx
                    break
    
    return columns_map


def parse_excel_file(file_path: str, expected_headers: List[str]) -> Tuple[pd.DataFrame, dict]:
    """
    Основной метод парсинга Excel файла.
    Возвращает DataFrame с данными и информацию о найденных колонках.
    """
    # Загружаем файл
    df = load_excel_file(file_path)
    
    # Находим начало таблицы
    header_row = find_table_start(df, expected_headers)
    
    # Используем найденную строку как заголовок
    df.columns = df.iloc[header_row].tolist()
    
    # Удаляем строки до заголовка
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    
    # Определяем колонки
    columns_info = detect_columns(df, 0)
    
    # Переименовываем колонки для удобства
    rename_map = {}
    if 'code' in columns_info:
        rename_map[df.columns[columns_info['code']]] = 'code'
    if 'status' in columns_info:
        rename_map[df.columns[columns_info['status']]] = 'status'
    if 'name' in columns_info:
        rename_map[df.columns[columns_info['name']]] = 'name'
    if 'special_price' in columns_info:
        rename_map[df.columns[columns_info['special_price']]] = 'special_price'
    if 'retail_price' in columns_info:
        rename_map[df.columns[columns_info['retail_price']]] = 'retail_price'
    
    df = df.rename(columns=rename_map)
    
    return df, columns_info
