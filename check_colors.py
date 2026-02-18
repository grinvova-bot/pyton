from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print('=== Строка 27 (распродажа) ===')
for j in range(1, 6):
    cell = ws.cell(row=27, column=j)
    fill = cell.fill.start_color.rgb if cell.fill.start_color.rgb else 'None'
    font_bold = cell.font.bold
    print(f'  {j}: {cell.value} [fill:{fill}, bold:{font_bold}]')

print('\n=== Проверка заголовка таблицы (строка 7) ===')
for j in range(1, 6):
    cell = ws.cell(row=7, column=j)
    fill = cell.fill.start_color.rgb if cell.fill.start_color.rgb else 'None'
    font_bold = cell.font.bold
    font_color = cell.font.color.rgb if cell.font.color and cell.font.color.rgb else 'None'
    print(f'  {j}: {cell.value} [fill:{fill}, bold:{font_bold}, color:{font_color}]')
