from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print('Строка 27 (распродажа):')
for j in range(1, 6):
    cell = ws.cell(row=27, column=j)
    fill = cell.fill.start_color.rgb if cell.fill.start_color.rgb else 'None'
    print(f'  {j}: {cell.value} [fill:{fill}]')

print('\nСтрока 28 (без распродажи):')
for j in range(1, 6):
    cell = ws.cell(row=28, column=j)
    fill = cell.fill.start_color.rgb if cell.fill.start_color.rgb else 'None'
    print(f'  {j}: {cell.value} [fill:{fill}]')
