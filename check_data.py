from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print("=== Все строки 7-35 ===")
for i in range(7, 35):
    row = []
    for j in range(1, 6):
        cell = ws.cell(row=i, column=j)
        row.append(str(cell.value)[:20] if cell.value else '')
    print(f"{i}: {row}")
