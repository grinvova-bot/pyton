from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print('=== Параметры страницы ===')
print(f'Ориентация: {ws.page_setup.orientation}')
print(f'Бумага: {ws.page_setup.paperSize}')
print(f'Масштаб: {ws.page_setup.scale}%')
print(f'fitToWidth: {ws.page_setup.fitToWidth}, fitToHeight: {ws.page_setup.fitToHeight}')
print(f'Поля: верх={ws.page_margins.top}, низ={ws.page_margins.bottom}, лево={ws.page_margins.left}, право={ws.page_margins.right}')

print('\n=== Сквозные строки ===')
print(f'print_title_rows: {ws.print_title_rows}')

print('\n=== Колонтитулы ===')
print(f'Верхний (центр): {ws.oddHeader.center.text if ws.oddHeader.center.text else "пусто"}')
print(f'Нижний (право): {ws.oddFooter.right.text if ws.oddFooter.right.text else "пусто"}')

print('\n=== Цвет рамок (строка 7, заголовок) ===')
for j in range(1, 6):
    cell = ws.cell(row=7, column=j)
    border_color = cell.border.left.color.rgb if cell.border.left.color and cell.border.left.color.rgb else 'None'
    print(f'  {j}: {cell.value} [border:{border_color}]')
