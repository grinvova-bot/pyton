"""
Модуль загрузки и парсинга Excel-файлов
"""
import pandas as pd
from openpyxl import load_workbook
from typing import Tuple, List, Optional
import re


class ExcelParser:
    """Парсер Excel-файлов со сложной структурой"""
    
    # Паттерны для распознавания колонок
    COLUMN_PATTERNS = {
        'code': [r'код\s*товара', r'код', r'артикул', r'№'],
        'status': [r'статус', r'распродажа', r'качество', r'срок', r'ограничен'],
        'name': [r'номенклатура', r'название', r'товар', r'описание'],
        'special_price': [r'спец\.?\s*цена', r'акция', r'специальная\s*цена'],
        'retail_price': [r'розничн.*цена', r'цена\s*розница', r'прайс\s*цена'],
    }
    
    # Паттерны для определения шапок и мусора
    HEADER_PATTERNS = [
        r'прайс[-\s]?лист',
        r'параметры',
        r'дата\s*отчета',
        r'ответственный',
        r'контакт',
        r'ООО\s*',
        r'^\d{1,2}\.\d{1,2}\.\d{4}',  # Даты
    ]
    
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.wb = load_workbook(file_path, data_only=True)
        self.ws = self.wb.active
        
    def detect_columns(self) -> dict:
        """
        Автоматическое определение индексов ключевых колонок.
        Возвращает словарь {column_type: column_index}
        """
        columns = {}
        
        # Читаем первые строки для поиска заголовков
        headers = []
        for row in self.ws.iter_rows(min_row=1, max_row=10, values_only=True):
            headers.append(row)
        
        # Ищем строку с заголовками таблицы
        header_row_idx = None
        for idx, row in enumerate(headers):
            row_text = ' '.join(str(c) for c in row if c)
            if 'код' in row_text.lower() and ('номенклатура' in row_text.lower() or 'товар' in row_text.lower()):
                header_row_idx = idx
                break
        
        if header_row_idx is not None:
            header_row = headers[header_row_idx]
            for col_idx, cell in enumerate(header_row):
                if cell is None:
                    continue
                cell_text = str(cell).lower().strip()
                
                for col_type, patterns in self.COLUMN_PATTERNS.items():
                    for pattern in patterns:
                        if re.search(pattern, cell_text):
                            columns[col_type] = col_idx
                            break
        
        return columns
    
    def read_raw_data(self) -> pd.DataFrame:
        """
        Чтение сырых данных из Excel.
        """
        # Находим строку с основными заголовками
        start_row = 1
        for row_idx, row in enumerate(self.ws.iter_rows(values_only=True), 1):
            row_text = ' '.join(str(c) for c in row if c)
            if 'код' in row_text.lower() and ('номенклатура' in row_text.lower() or 'товар' in row_text.lower()):
                start_row = row_idx + 1  # +1 потому что следующая строка после заголовков
                break
        
        # Читаем данные начиная с найденной строки
        df = pd.read_excel(
            self.file_path,
            header=None,
            skiprows=start_row,
            engine='openpyxl'
        )
        
        return df
    
    def get_all_rows(self) -> List[tuple]:
        """Получить все строки листа как кортежи"""
        rows = []
        for row in self.ws.iter_rows(values_only=True):
            rows.append(row)
        return rows
    
    def close(self):
        """Закрытие книги"""
        self.wb.close()
