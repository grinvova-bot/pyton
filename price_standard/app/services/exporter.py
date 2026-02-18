"""
Модуль экспорта и шаблонизации
Оформление Excel-файла в корпоративном стиле
"""
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.header_footer import HeaderFooter
import pandas as pd
from datetime import datetime
from typing import Optional
import numpy as np


class ExcelExporter:
    """Экспорт данных в Excel с корпоративным стилем"""

    # Стили оформления
    HEADER_FONT = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='444444', end_color='444444', fill_type='solid')
    DATA_FONT = Font(name='Arial', size=10)
    PRICE_FONT = Font(name='Arial', size=10)
    PRICE_FORMAT = '# ##0'
    SALE_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    SALE_STATUS_FONT = Font(name='Arial', size=10, bold=True, color='FF0000')

    # Стили для заголовков разделов (категорий товаров)
    CATEGORY_FONT = Font(name='Arial', size=11, bold=True)
    CATEGORY_FILL = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')  # Бледно-голубой

    # Стили для шапки документа
    DOC_HEADER_FONT = Font(name='Arial', size=14, bold=True)
    DOC_HEADER_FONT_SMALL = Font(name='Arial', size=10)

    THIN_BORDER = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    
    def __init__(self, df: pd.DataFrame, column_mapping: dict):
        self.df = df.reset_index(drop=True)
        self.column_mapping = column_mapping
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Прайс-лист"
        
    def _clean_value(self, value):
        """Очистка значения для записи в Excel"""
        if value is None or pd.isna(value):
            return None
        if isinstance(value, (np.integer, np.int64)):
            return int(value)
        if isinstance(value, (np.floating, np.float64)):
            return float(value)
        if isinstance(value, (list, dict)):
            return str(value)
        return value
        
    def _apply_header_style(self, row: int, num_cols: int):
        for col in range(1, num_cols + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER
            
    def _apply_data_styles(self, start_row: int, num_cols: int, price_cols: list):
        for row_idx in range(start_row, self.ws.max_row + 1):
            for col_idx in range(1, num_cols + 1):
                cell = self.ws.cell(row=row_idx, column=col_idx)
                cell.border = self.THIN_BORDER
                cell.font = self.DATA_FONT
                
                # Формат для цен
                if col_idx in price_cols:
                    cell.number_format = self.PRICE_FORMAT
                    cell.alignment = Alignment(horizontal='right', vertical='center')
    
    def _highlight_sale_rows(self, start_row: int, status_col_idx: int, num_cols: int):
        for row_idx in range(start_row, self.ws.max_row + 1):
            status_cell = self.ws.cell(row=row_idx, column=status_col_idx)
            if status_cell.value == "РАСПРОДАЖА":
                for col in range(1, num_cols + 1):
                    cell = self.ws.cell(row=row_idx, column=col)
                    cell.fill = self.SALE_FILL
                status_cell.font = self.SALE_STATUS_FONT
    
    def _is_table_header(self, row_data: list) -> bool:
        """Проверка, является ли строка заголовком таблицы (не нужна в выводе)"""
        text = ' '.join(str(v) for v in row_data if v).lower()
        header_words = ['номенклатура', 'код товара', 'цена', 'качество', 'спец', 'розничн']
        matches = sum(1 for word in header_words if word in text)
        return matches >= 2

    def _is_category_header(self, row_data: list) -> bool:
        """
        Проверка, является ли строка заголовком раздела.
        Заголовок раздела - строка где:
        - В колонке кода нет цифр (или пусто)
        - В колонке номенклатуры текст без цен
        - Обычно это названия типа "AND ACOMIX", "BOSTIK" и т.д.
        """
        if len(row_data) < 3:
            return False
            
        code_val = row_data[0] if len(row_data) > 0 else None
        status_val = row_data[1] if len(row_data) > 1 else None
        name_val = row_data[2] if len(row_data) > 2 else None
        
        # Если код товара пустой или текстовый (без цифр)
        code_is_text = False
        if code_val is None or (isinstance(code_val, str) and not code_val.strip()):
            code_is_text = True
        elif isinstance(code_val, str):
            code_str = code_val.strip()
            # Если нет цифр и длина небольшая - возможно это категория
            if not any(c.isdigit() for c in code_str) and len(code_str) < 50:
                code_is_text = True
        
        # Проверяем, что статус пустой (у категорий нет статуса)
        status_is_empty = status_val is None or (isinstance(status_val, str) and not status_val.strip())
        
        # Если код текстовый и статус пустой - это категория
        if code_is_text and status_is_empty:
            return True
        
        return False

    def _apply_category_style(self, row: int, num_cols: int):
        """Применение стиля заголовка раздела"""
        for col in range(1, num_cols + 1):
            cell = self.ws.cell(row=row, column=col)
            cell.font = self.CATEGORY_FONT
            cell.fill = self.CATEGORY_FILL
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = self.THIN_BORDER

    def _add_document_header(self):
        """Добавление шапки документа (6 строк по ТЗ №2)"""
        from datetime import datetime
        
        # Форматируем текущую дату на русском языке
        months = {
            1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля',
            5: 'мая', 6: 'июня', 7: 'июля', 8: 'августа',
            9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'
        }
        now = datetime.now()
        date_str = f"дата: {now.day} {months[now.month]} {now.year} г."
        
        # Вставляем 6 строк в начало (по ТЗ: строки 1-6)
        self.ws.insert_rows(1, amount=6)
        
        # Объединяем ячейки для заголовка
        last_col = self.ws.max_column
        last_col_letter = get_column_letter(last_col)
        
        # Строка 1: Прайс-лист ООО "АЛЬТ-Икс"
        self.ws.merge_cells(f'A1:{last_col_letter}1')
        cell = self.ws.cell(row=1, column=1)
        cell.value = 'Прайс-лист ООО "АЛЬТ-Икс"'
        cell.font = self.DOC_HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Строка 2: пустая
        self.ws.row_dimensions[2].height = 15
        
        # Строка 3: дата
        self.ws.cell(row=3, column=1, value=date_str)
        self.ws.cell(row=3, column=1).font = self.DOC_HEADER_FONT_SMALL
        
        # Строка 4: пустая
        self.ws.row_dimensions[4].height = 15
        
        # Строка 5: ответственный
        self.ws.cell(row=5, column=1, value='Ответственный: Сидорова О.О.')
        self.ws.cell(row=5, column=1).font = self.DOC_HEADER_FONT_SMALL
        
        # Строка 6: будет заголовок таблицы (сдвигается)

    def _setup_page_layout(self):
        """Настройка параметров страницы и колонтитулов (по ТЗ №2)"""
        # Параметры страницы
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE
        self.ws.page_setup.paperSize = self.ws.PAPERSIZE_A4
        
        # Поля (см)
        self.ws.page_margins.top = 2.0
        self.ws.page_margins.bottom = 1.5
        self.ws.page_margins.left = 1.5
        self.ws.page_margins.right = 1.5
        self.ws.page_margins.header = 1.0
        self.ws.page_margins.footer = 1.0
        
        # Масштаб - вписать по ширине
        self.ws.page_setup.fitToWidth = 1
        self.ws.page_setup.fitToHeight = None
        
        # Колонтитулы
        # Верхний: дата по центру
        self.ws.oddHeader.center.text = '&D'  # Дата
        self.ws.oddHeader.center.size = 9
        
        # Нижний: номер страницы справа
        self.ws.oddFooter.right.text = 'Страница &P из &N'
        self.ws.oddFooter.right.size = 9

    def _auto_adjust_column_widths(self, num_cols: int):
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row_idx in range(1, self.ws.max_row + 1):
                cell_value = self.ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def export(self, output_path: Optional[str] = None) -> str:
        # Порядок колонок для вывода
        column_order = ['code', 'status', 'name', 'special_price', 'retail_price']
        column_names = {
            'code': 'Код товара',
            'status': 'Статус',
            'name': 'Номенклатура',
            'special_price': 'Спец. цена',
            'retail_price': 'Розничная цена'
        }

        # Формируем выходные данные
        output_data = []
        headers = []
        price_cols = []  # Индексы колонок с ценами (1-based)

        for idx, col_type in enumerate(column_order, 1):
            if col_type in self.column_mapping:
                col_idx = self.column_mapping[col_type]
                if col_idx < len(self.df.columns):
                    headers.append(column_names.get(col_type, col_type))
                    col_data = []
                    for val in self.df.iloc[:, col_idx]:
                        col_data.append(self._clean_value(val))
                    output_data.append(col_data)
                    if 'price' in col_type:
                        price_cols.append(idx)

        num_cols = len(output_data)
        num_rows = len(output_data[0]) if output_data else 0

        # Запись заголовков
        for c_idx, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=c_idx, value=header)

        # Запись данных (пропускаем строки-заголовки таблиц)
        actual_row = 2  # Реальный номер строки в Excel (после шапки будет сдвиг)
        category_rows = []  # Номера строк с категориями
        sale_check_rows = []  # Строки для проверки на распродажу
        
        for r_idx in range(num_rows):
            row_values = [output_data[c_idx][r_idx] if r_idx < len(output_data[c_idx]) else None for c_idx in range(num_cols)]
            
            # Пропускаем строки, которые выглядят как заголовки таблицы
            if self._is_table_header(row_values):
                continue
            
            # Записываем данные
            for c_idx in range(num_cols):
                value = output_data[c_idx][r_idx] if r_idx < len(output_data[c_idx]) else None
                self.ws.cell(row=actual_row, column=c_idx + 1, value=value)
            
            # Проверяем, является ли строка заголовком раздела
            if self._is_category_header(row_values):
                category_rows.append(actual_row)
            
            # Сохраняем для проверки на распродажу
            sale_check_rows.append((actual_row, row_values))
            
            actual_row += 1
        
        # Реальное количество строк данных
        data_rows = actual_row - 2

        # Стили
        self._apply_header_style(1, num_cols)
        self._apply_data_styles(2, num_cols, price_cols)

        # Выделение заголовков разделов (категорий)
        for row_num in category_rows:
            self._apply_category_style(row_num, num_cols)

        # Выделение распродажи
        status_col_idx = 2 if 'status' in self.column_mapping else 1
        for row_num, row_values in sale_check_rows:
            status_val = row_values[1] if len(row_values) > 1 else None
            if status_val == "РАСПРОДАЖА":
                for col in range(1, num_cols + 1):
                    cell = self.ws.cell(row=row_num, column=col)
                    cell.fill = self.SALE_FILL
                status_cell = self.ws.cell(row=row_num, column=status_col_idx)
                status_cell.font = self.SALE_STATUS_FONT

        # Добавление шапки документа
        self._add_document_header()

        # Автоподбор ширины
        self._auto_adjust_column_widths(num_cols)

        # Автофильтр (с учетом сдвига на 6 строк шапки, заголовок на строке 7)
        if num_cols > 0 and data_rows > 0:
            last_col = get_column_letter(num_cols)
            self.ws.auto_filter.ref = f"A7:{last_col}{data_rows + 7}"

        # Имя файла
        if output_path is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"Прайс_Стандарт_{timestamp}.xlsx"

        self.wb.save(output_path)
        self.wb.close()

        return output_path


def generate_output_filename() -> str:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"Прайс_Стандарт_{timestamp}.xlsx"
