from openpyxl import load_workbook
import pandas as pd

print('=== Data.xlsx ===')
wb = load_workbook('C:/Users/Grintsov/Pyton/Griin/Data.xlsx', data_only=True)
print('Листы:', wb.sheetnames)
ws = wb.active
print('Активный лист:', ws.title)

print('\nПервые 30 строк:')
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
    print(f'{i}: {row}')

print('\n=== Temlate.xlsx ===')
wb2 = load_workbook('C:/Users/Grintsov/Pyton/Griin/Temlate.xlsx', data_only=True)
print('Листы:', wb2.sheetnames)
ws2 = wb2.active
print('Активный лист:', ws2.title)

print('\nШаблон (первые 10 строк):')
for i, row in enumerate(ws2.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    print(f'{i}: {row}')

# Чтение через pandas
print('\n=== Pandas анализ Data.xlsx ===')
df = pd.read_excel('C:/Users/Grintsov/Pyton/Griin/Data.xlsx')
print(f'Размер: {df.shape}')
print(f'Колонки: {df.columns.tolist()}')
print('\nПервые 10 строк:')
print(df.head(10).to_string())
