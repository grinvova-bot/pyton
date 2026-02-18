from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print('=== Структура файла ===')
for i in range(1, 12):
    row = []
    for j in range(1, 6):
        cell = ws.cell(row=i, column=j)
        value = str(cell.value)[:25] if cell.value else ''
        row.append(value)
    print(f'{i}: {row}')
