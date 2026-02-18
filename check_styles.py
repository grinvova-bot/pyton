from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print("=== Первые 15 строк ===")
for i in range(1, 16):
    row_data = []
    for j in range(1, 6):
        cell = ws.cell(row=i, column=j)
        fill_color = cell.fill.start_color.rgb if cell.fill.start_color.rgb else 'None'
        value = str(cell.value)[:30] if cell.value else ''
        row_data.append(f"{value} [fill:{fill_color}]")
    print(f"Строка {i}: {' | '.join(row_data[:3])}")

print("\n=== Проверка шапки ===")
print(f"Строка 1: {ws.cell(row=1, column=1).value}")
print(f"Строка 3: {ws.cell(row=3, column=1).value}")
print(f"Строка 4: {ws.cell(row=4, column=1).value}")
