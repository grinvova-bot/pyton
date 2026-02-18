from openpyxl import load_workbook
from datetime import datetime

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print('Строка 1:', ws.cell(row=1, column=1).value)
print('Строка 3:', ws.cell(row=3, column=1).value)
print('Строка 4:', ws.cell(row=4, column=1).value)
print()
print('Ожидаемая дата:', datetime.now().strftime('%d %B %Y'))
