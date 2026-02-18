from openpyxl import load_workbook

wb = load_workbook('c:/Users/Grintsov/Pyton/price_standard/output/test_result.xlsx')
ws = wb.active

print('=== Структура файла (первые 12 строк) ===')
for i in range(1, 13):
    row = []
    for j in range(1, 6):
        cell = ws.cell(row=i, column=j)
        value = str(cell.value)[:25] if cell.value else ''
        fill = cell.fill.start_color.rgb if cell.fill.start_color.rgb else 'None'
        row.append(f'{value} [{fill}]')
    print(f'{i}: {row[:2]}')  # Первые 2 колонки для краткости

print('\n=== Параметры страницы ===')
print(f'Ориентация: {ws.page_setup.orientation}')
print(f'Поля: верх={ws.page_margins.top}, низ={ws.page_margins.bottom}, лево={ws.page_margins.left}, право={ws.page_margins.right}')
print(f'Масштаб: fitToWidth={ws.page_setup.fitToWidth}, fitToHeight={ws.page_setup.fitToHeight}')

print('\n=== Колонтитулы ===')
print(f'Верхний (центр): {ws.oddHeader.center.text if ws.oddHeader.center.text else "пусто"}')
print(f'Нижний (право): {ws.oddFooter.right.text if ws.oddFooter.right.text else "пусто"}')
